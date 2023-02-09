from PyPDF2 import PdfReader
import zipfile, os
from os.path import basename
from openpyxl import load_workbook
import csv


path_into_zip = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
path_out_zip = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
files_dir = os.listdir(path_into_zip)
path_zip = os.path.join(path_out_zip, "test.zip")


def test_create_archive():
    with zipfile.ZipFile(path_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in files_dir:
            add_file = os.path.join(path_into_zip, file)
            zf.write(add_file, basename(add_file))


def test_read_from_pdf():
    with zipfile.ZipFile(path_zip) as zf:
        pdf_file = zf.extract("test_pdf.pdf")
        reader = PdfReader(pdf_file)
        page = reader.pages[0]
        text = page.extract_text()
        assert 'Star Wars' in text, f"Expected result: {'Star Wars'}; actual result: {text}"

        os.remove(pdf_file)


def test_read_from_xlsx_file():

    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("test_xlsx.xlsx")
        workbook = load_workbook(cf)
        sheet = workbook.active
        film = sheet.cell(row=2, column=1).value
        assert film == 'Star Wars', f"Expected result: {'Star Wars'}; actual result: {film}"

        os.remove("test_xlsx.xlsx")


def test_read_from_csv_file():

    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("test_csv.csv")
        with open(cf) as csvfile:
            csvfile = csv.reader(csvfile)
            new_list = []
            for row in csvfile:
                new_list.append(row)
            assert len(new_list) == 7
            assert ['Star Wars', '1978'] in new_list

        os.remove("test_csv.csv")